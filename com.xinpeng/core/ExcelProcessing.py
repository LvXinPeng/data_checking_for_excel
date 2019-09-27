import pandas as pd
from openpyxl import load_workbook
import openpyxl

# import xlsxwriter
df = pd.read_excel("D:/record-part.xlsx")
# df = pd.read_excel("C:\\Users\\xinplv\\Downloads\\[CS_BI_Report_Interivew_&_Evaluation]_For_人员能力模型\\原始数据：认证记录.xlsx")

# df_new = pd.read_excel("D:/record-new.xlsx")
# 输出 所有列名，加索引[2] 可定位某一列名；
# print(range(len(df.columns)))
# print(df['经销商代码'])

# 输出整列的匹配数；
# print(df.groupby("三级机械技师").size()['100%'])
agent_code = df['经销商代码']
agent_code_unique = agent_code.unique()
# print(len(agent_code_unique))

# book = load_workbook('D:/record-new.xlsx')
writer = pd.ExcelWriter(path="D:/record-new.xlsx", mode='w', engine='openpyxl')
# writer.book = book
# writer.sheets = {ws.title: ws for ws in book.worksheets}
# mx_row = writer.sheets['Sheet1'].max_row
# print(mx_row)
# row_data = []
rd = []
for idx in range(len(agent_code_unique)):
    row_data = []
    row_data.append(agent_code_unique[idx])
    # print(row_data)
    for index in range(len(df.columns) - 5):
        # 判断值 ('ASA', '100%') 输出值为True or False
        isExist = (agent_code_unique[idx], '100%') in df.groupby([agent_code, df.columns[index + 5]]).groups
        if isExist:
            size_col = df.groupby([agent_code, df.columns[index + 5]]).size()[agent_code_unique[idx], '100%']
            row_data.append(size_col)
            # print(str(agent_code_unique[idx]) + "-" + str(df.columns[index + 5]) + ": " + str(size_col))
        else:
            # 如果在判断 不存在('ASA', '100%')这样类似的组合 即认定为不存在符合要求的技师
            # print(str(agent_code_unique[idx]) + "-" + str(df.columns[index + 5]) + ": " + str(0))
            row_data.append(0)
    rd.append(row_data)
col = ['一级钣金技师	', '一级技师', '沃尔沃入职证书 Volvo Onboard Certification', '二级技师', '二级钣金技师', '三级公共技师', '三级钣金技师', '一级专属服务顾问',
       '三级电气技师', '三级机械技师', '三级空调技师', '四级技师', '二级专属服务顾问', '一级专属服务技师', '客户服务经理', '车间管理', '保修专员', '高级保修专员', '一级配件专员',
       '二级配件专员', '经销商内训师', '服务管家', '客户关系经理', '混合动力技师']
frame = pd.DataFrame(rd)
# print(frame.dtypes)
frame.to_excel(writer, index=False, header=True)
writer.save()
print(rd)
writer.close()

print('===========================')
# print(rd)


# df_new.to_excel()
# groupby = df.groupby([agent_code, df.columns[10]])
# print((agent_code[1], '100%') in groupby.groups)
# print(groupby.groups)
# print(groupby.size()[agent_code_unique[2], "100%"])
